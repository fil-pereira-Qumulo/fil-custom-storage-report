"""
storage_report.py

Collects storage metrics from the Qumulo API, writes them to InfluxDB,
and emails a monthly Excel report via SMTP.  Designed to run as a service.

Requirements:
    pip install requests influxdb-client openpyxl

Python >= 3.11  (uses built-in tomllib; for 3.10 and below: pip install tomli)
"""

import argparse
import base64
import contextlib
import logging
import os
import signal
import smtplib
import stat
import sys
import tempfile
import time
import tomllib
import unittest
import warnings
from dataclasses import dataclass, field
from datetime import datetime, timezone
from email.message import EmailMessage
from io import BytesIO
from pathlib import Path
from typing import Any
from unittest.mock import MagicMock
from urllib.parse import quote, urlparse

import requests
import urllib3
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from influxdb_client import InfluxDBClient, Point, WritePrecision
from influxdb_client.client.write_api import SYNCHRONOUS

# ---------------------------------------------------------------------------
# Logging
# ---------------------------------------------------------------------------

# basicConfig is called in __main__ so that importing this module (e.g. in tests
# or by another program) does not reconfigure the root logger unexpectedly.
log = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# Configuration
# ---------------------------------------------------------------------------

CONFIG_PATH  = Path(__file__).parent / "config.toml"
_STATE_PATH  = Path(__file__).parent / ".report_state"


@dataclass
class QumuloConfig:
    host: str
    token: str
    cluster_name: str
    port: int = 8000
    tenant_hosts: dict[str, str] = field(default_factory=dict)  # tenant name → FQDN


@dataclass
class InfluxConfig:
    url: str
    token: str
    org: str
    bucket: str


@dataclass
class SmtpConfig:
    host: str
    port: int
    from_addr: str
    to_addrs: list[str]
    use_tls: bool = True
    username: str | None = None
    password: str | None = None


@dataclass
class ReportConfig:
    schedule_day: int
    schedule_hour: int
    collection_interval_minutes: int = 5
    title: str = "Monthly Storage Report"


@dataclass
class AppConfig:
    clusters: list[QumuloConfig]
    influxdb: InfluxConfig
    smtp: SmtpConfig
    report: ReportConfig


def _check_config_permissions(path: Path) -> None:
    """On Unix systems, warn if the config file is readable by group or others."""
    if sys.platform == "win32":
        return
    mode = path.stat().st_mode
    if mode & (stat.S_IRGRP | stat.S_IROTH):
        log.warning(
            "Config file %s is readable by group or others — it contains tokens. "
            "Restrict permissions with: chmod 600 %s", path, path,
        )


def _validate_config(cfg: AppConfig) -> None:
    """Raise ValueError if any config values are outside their valid ranges."""
    errors: list[str] = []
    if not cfg.clusters:
        errors.append("at least one [[clusters]] entry is required")
    seen_names: set[str] = set()
    for cluster in cfg.clusters:
        if not cluster.host:
            errors.append(f"cluster with name '{cluster.cluster_name}' has an empty host")
        if not cluster.token:
            errors.append(f"cluster with host '{cluster.host}' has an empty token")
        if not cluster.cluster_name:
            errors.append(f"cluster with host '{cluster.host}' has an empty cluster_name")
        if cluster.cluster_name in seen_names:
            errors.append(f"duplicate cluster_name '{cluster.cluster_name}' — each cluster must have a unique name")
        seen_names.add(cluster.cluster_name)
    if not cfg.influxdb.url:
        errors.append("influxdb.url must not be empty")
    if not cfg.smtp.host:
        errors.append("smtp.host must not be empty")
    if not 1 <= cfg.smtp.port <= 65535:
        errors.append(f"smtp.port must be 1–65535, got {cfg.smtp.port}")
    if bool(cfg.smtp.username) != bool(cfg.smtp.password):
        errors.append("smtp.username and smtp.password must both be set or both be omitted")
    if not 1 <= cfg.report.schedule_day <= 28:
        errors.append(f"report.schedule_day must be 1–28, got {cfg.report.schedule_day}")
    if not 0 <= cfg.report.schedule_hour <= 23:
        errors.append(f"report.schedule_hour must be 0–23, got {cfg.report.schedule_hour}")
    if cfg.report.collection_interval_minutes < 1:
        errors.append(
            f"report.collection_interval_minutes must be >= 1, "
            f"got {cfg.report.collection_interval_minutes}"
        )
    if not cfg.smtp.to_addrs:
        errors.append("smtp.to_addrs must contain at least one recipient")
    if errors:
        raise ValueError("Invalid configuration:\n" + "\n".join(f"  - {e}" for e in errors))


def load_config(path: Path = CONFIG_PATH) -> AppConfig:
    _check_config_permissions(path)
    with open(path, "rb") as f:
        raw = tomllib.load(f)

    i = raw["influxdb"]
    s = raw["smtp"]
    r = raw["report"]

    clusters = [
        QumuloConfig(
            host=c["host"],
            port=c.get("port", 8000),
            token=c["token"],
            cluster_name=c["cluster_name"],
            tenant_hosts=c.get("tenant_hosts", {}),
        )
        for c in raw.get("clusters", [])
    ]

    cfg = AppConfig(
        clusters=clusters,
        influxdb=InfluxConfig(
            url=i["url"],
            token=i["token"],
            org=i["org"],
            bucket=i["bucket"],
        ),
        smtp=SmtpConfig(
            host=s["host"],
            port=s["port"],
            use_tls=s.get("use_tls", True),
            username=s.get("username"),
            password=s.get("password"),
            from_addr=s["from_addr"],
            to_addrs=s["to_addrs"],
        ),
        report=ReportConfig(
            schedule_day=r["schedule_day"],
            schedule_hour=r["schedule_hour"],
            collection_interval_minutes=r.get("collection_interval_minutes", 5),
            title=r.get("title", "Monthly Storage Report"),
        ),
    )
    _validate_config(cfg)
    return cfg

# ---------------------------------------------------------------------------
# Qumulo API client
# ---------------------------------------------------------------------------

_MAX_API_PAGES = 1000  # safety cap to prevent infinite loops on misbehaving pagination


class QumuloClient:
    """Thin wrapper around the Qumulo REST API."""

    def __init__(self, cfg: QumuloConfig) -> None:
        self.base_url = f"https://{cfg.host}:{cfg.port}"
        self.headers = {
            "Authorization": f"Bearer {cfg.token}",
            "Content-Type": "application/json",
        }
        self.cluster_name = cfg.cluster_name

    def _get(self, path: str, params: dict | None = None) -> Any:
        url = f"{self.base_url}{path}"
        # catch_warnings() modifies global warning-filter state and is not thread-safe.
        # Safe here because the service runs single-threaded.
        with warnings.catch_warnings():
            warnings.simplefilter("ignore", urllib3.exceptions.InsecureRequestWarning)
            resp = requests.get(url, headers=self.headers, params=params, verify=False, timeout=30)
        resp.raise_for_status()
        return resp.json()

    def _get_all_pages(self, path: str, list_key: str, params: dict | None = None) -> list[dict]:
        """Fetch every page of a paged Qumulo endpoint, following paging.next until exhausted."""
        results: list[dict] = []
        current_path: str | None = path
        current_params = params
        pages_fetched = 0
        while current_path:
            if pages_fetched >= _MAX_API_PAGES:
                log.warning(
                    "Pagination limit (%d pages) reached for %s — results may be incomplete",
                    _MAX_API_PAGES, path,
                )
                break
            data = self._get(current_path, current_params)
            results.extend(data.get(list_key, []))
            pages_fetched += 1
            raw_next = (data.get("paging") or {}).get("next") or ""
            if raw_next:
                parsed = urlparse(raw_next)
                qs = f"?{parsed.query}" if parsed.query else ""
                # If paging.next has no path (e.g. bare "?after=abc"), fall back
                # to the original endpoint path so the URL stays well-formed.
                current_path = (parsed.path or path) + qs
                current_params = None   # next URL already encodes all params
            else:
                current_path = None
        return results

    def get_nfs_exports(self) -> list[dict]:
        return self._get_all_pages("/v3/nfs/exports/", "entries")

    def get_smb_shares(self) -> list[dict]:
        return self._get_all_pages(
            "/v3/smb/shares/", "entries", params={"populate-trustee-names": "false"}
        )

    def get_quotas(self) -> list[dict]:
        return self._get_all_pages("/v1/files/quotas/status/", "quotas")

    def get_tenants(self) -> dict[int, str]:
        """Return a mapping of tenant_id → tenant_name."""
        entries = self._get_all_pages("/v1/multitenancy/tenants/", "entries")
        result: dict[int, str] = {}
        for entry in entries:
            tid  = entry.get("id")
            name = entry.get("name", "")
            if tid is None:
                log.warning("get_tenants: entry missing 'id' field — skipping: %s", entry)
                continue
            try:
                result[int(tid)] = name
            except (TypeError, ValueError):
                log.warning("get_tenants: non-integer 'id' value %r — skipping", tid)
                continue
        return result

    def get_file_metadata(self, path: str) -> dict[str, str]:
        """
        Return generic user-metadata for *path* as a plain {key: value} dict.
        Values are base64-decoded from the API response.

        All HTTP errors (404 = no metadata, 4xx/5xx = API problem) return an
        empty dict with a warning rather than aborting the entire cluster fetch.
        Pagination is followed so more than 100 metadata entries are handled.
        """
        encoded = quote(path.lstrip("/"), safe="")
        try:
            entries = self._get_all_pages(
                f"/v1/files/%2F{encoded}/user-metadata/generic/",
                "entries",
            )
        except requests.RequestException as exc:
            status = (
                exc.response.status_code
                if isinstance(exc, requests.HTTPError) and exc.response is not None
                else type(exc).__name__
            )
            log.warning("get_file_metadata: %s for path '%s' — returning empty metadata",
                        status, path)
            return {}

        result: dict[str, str] = {}
        for entry in entries:
            key = entry.get("key", "")
            raw_value = entry.get("value", "")
            try:
                result[key] = base64.b64decode(raw_value).decode("utf-8")
            except Exception:
                result[key] = raw_value  # fall back to raw if decode fails
        return result

# ---------------------------------------------------------------------------
# InfluxDB writer
# ---------------------------------------------------------------------------

class InfluxWriter:
    """Writes storage metrics to InfluxDB."""

    def __init__(self, cfg: InfluxConfig, cluster_name: str) -> None:
        self.client = InfluxDBClient(url=cfg.url, token=cfg.token, org=cfg.org)
        self.write_api = self.client.write_api(write_options=SYNCHRONOUS)
        self.bucket = cfg.bucket
        self.org = cfg.org
        self.cluster = cluster_name

    def write_quota_metrics(self, quotas: list[dict]) -> None:
        # The quota API (/v1/files/quotas/status/) does not return tenant_id,
        # so quota points are tagged by cluster and path only.
        points = []
        for q in quotas:
            path  = q.get("path", "unknown")
            limit = int(q.get("limit") or 0)
            used  = int(q.get("capacity_usage") or 0)
            pct   = (used / limit * 100) if limit else 0.0

            point = (
                Point("storage_quota")
                .tag("cluster", self.cluster)
                .tag("path", path)
                .field("limit_bytes", limit)
                .field("used_bytes", used)
                .field("used_pct", round(pct, 2))
                .time(datetime.now(timezone.utc), WritePrecision.S)
            )
            points.append(point)

        if not points:
            log.info("InfluxDB: no quota points to write")
            return
        self.write_api.write(bucket=self.bucket, org=self.org, record=points)
        log.info("InfluxDB: wrote %d quota points", len(points))

    def write_smb_share_count(self, shares: list[dict], tenant_names: dict[int, str]) -> None:
        # Write one point per tenant with that tenant's share count
        counts: dict[int, int] = {}
        for s in shares:
            tid = int(s.get("tenant_id") or 0)
            counts[tid] = counts.get(tid, 0) + 1

        points = [
            Point("storage_smb")
            .tag("cluster", self.cluster)
            .tag("tenant", tenant_names.get(tenant_id, str(tenant_id)))
            .field("share_count", count)
            .time(datetime.now(timezone.utc), WritePrecision.S)
            for tenant_id, count in counts.items()
        ]
        if not points:
            log.info("InfluxDB: no SMB share count points to write")
            return
        self.write_api.write(bucket=self.bucket, org=self.org, record=points)
        log.info("InfluxDB: wrote SMB share counts across %d tenant(s)", len(counts))

    def write_nfs_export_count(self, exports: list[dict], tenant_names: dict[int, str]) -> None:
        # Write one point per tenant with that tenant's export count
        counts: dict[int, int] = {}
        for e in exports:
            tid = int(e.get("tenant_id") or 0)
            counts[tid] = counts.get(tid, 0) + 1

        points = [
            Point("storage_nfs")
            .tag("cluster", self.cluster)
            .tag("tenant", tenant_names.get(tenant_id, str(tenant_id)))
            .field("export_count", count)
            .time(datetime.now(timezone.utc), WritePrecision.S)
            for tenant_id, count in counts.items()
        ]
        if not points:
            log.info("InfluxDB: no NFS export count points to write")
            return
        self.write_api.write(bucket=self.bucket, org=self.org, record=points)
        log.info("InfluxDB: wrote NFS export counts across %d tenant(s)", len(counts))

    def write_share_details(self, rows: list[dict]) -> None:
        """Write one point per SMB share / NFS export with all Excel columns as fields."""
        points = []
        for row in rows:
            point = (
                Point("storage_share_detail")
                .tag("cluster",    self.cluster)
                .tag("tenant",     row.get("tenant_name", ""))
                .tag("share_path", row.get("share_path", ""))
                .field("type",         row.get("type", ""))
                .field("fs_path",      row.get("fs_path", ""))
                .field("quota_limit",  row.get("quota_limit", ""))
                .field("quota_used",   row.get("quota_used", ""))
                .field("owner",        row.get("owner", ""))
                .field("team",         row.get("team", ""))
                .field("snow_request", row.get("snow", ""))
                .field("speedtype",    row.get("speedtype", ""))
                .time(datetime.now(timezone.utc), WritePrecision.S)
            )
            points.append(point)

        if not points:
            log.info("InfluxDB: no share/export detail points to write")
            return
        self.write_api.write(bucket=self.bucket, org=self.org, record=points)
        log.info("InfluxDB: wrote %d share/export detail points", len(points))

    def close(self) -> None:
        self.client.close()

# ---------------------------------------------------------------------------
# Report data assembly
# ---------------------------------------------------------------------------

def _bytes_to_human(b: int) -> str:
    """Convert a byte count to a human-readable string (TB/GB/MB)."""
    if b < 0:
        log.warning("_bytes_to_human received negative value %d — clamping to 0", b)
        b = 0
    for unit, threshold in [("PB", 10**15), ("TB", 10**12), ("GB", 10**9), ("MB", 10**6)]:
        if b >= threshold:
            return f"{b / threshold:.2f}{unit}"
    return f"{b}B"


def _tenant_fqdn(tenant_id: int, tenant_names: dict[int, str], tenant_hosts: dict[str, str]) -> str:
    """Resolve tenant_id → tenant name → FQDN configured by the user."""
    name = tenant_names.get(tenant_id, "")
    return tenant_hosts.get(name, "unknown-tenant")


def assemble_report_rows(
    qumulo: QumuloClient,
    tenant_names: dict[int, str],
    tenant_hosts: dict[str, str],
    smb_shares: list[dict],
    nfs_exports: list[dict],
    quotas: list[dict],
) -> list[dict]:
    """
    Build one dict per share/export with all columns needed for the report.
    Quota data is joined by filesystem path.  File metadata is fetched per path.
    """
    # Build quota lookup: normalised path → {limit, used}
    quota_map: dict[str, dict] = {}
    for q in quotas:
        path = q.get("path", "").rstrip("/") or "/"
        if path in quota_map:
            log.warning("Duplicate quota path '%s' — keeping last entry", path)
        quota_map[path] = q

    rows: list[dict] = []
    metadata_cache: dict[str, dict[str, str]] = {}

    def _metadata(path: str) -> dict[str, str]:
        if path not in metadata_cache:
            metadata_cache[path] = qumulo.get_file_metadata(path)
        return metadata_cache[path]

    for share in smb_shares:
        fs_path = share.get("fs_path", "").rstrip("/") or "/"
        if fs_path == "/":
            log.debug("Skipping SMB share id=%s — fs_path is root '/'", share.get("id", "?"))
            continue
        tenant_id   = int(share.get("tenant_id") or 0)
        tenant_name = tenant_names.get(tenant_id, str(tenant_id))
        if tenant_name not in tenant_hosts:
            log.debug("Skipping SMB share id=%s — tenant %r not in tenant_hosts", share.get("id", "?"), tenant_name)
            continue
        data_host = _tenant_fqdn(tenant_id, tenant_names, tenant_hosts)
        share_name = share.get("share_name", "")
        if not share_name:
            log.warning("SMB share id=%s has no share_name — path will be incomplete", share.get("id", "?"))

        quota = quota_map.get(fs_path)
        limit = int(quota.get("limit") or 0)          if quota is not None else None
        used  = int(quota.get("capacity_usage") or 0) if quota is not None else None

        metadata = _metadata(fs_path)

        rows.append({
            "cluster_name": qumulo.cluster_name,
            "share_path":   f"\\\\{data_host}\\{share_name}",
            "type":         "SMB",
            "fs_path":      fs_path,
            "quota_limit":  _bytes_to_human(limit) if limit is not None else "",
            "quota_used":   _bytes_to_human(used)  if used  is not None else "",
            "owner":        metadata.get("Owner", ""),
            "team":         metadata.get("Team", ""),
            "snow":         metadata.get("SNOW Request", ""),
            "speedtype":    metadata.get("Speedtype", ""),
            "tenant_id":    tenant_id,
            "tenant_name":  tenant_names.get(tenant_id, str(tenant_id)),
        })

    for export in nfs_exports:
        fs_path = export.get("fs_path", "").rstrip("/") or "/"
        if fs_path == "/":
            log.debug("Skipping NFS export id=%s — fs_path is root '/'", export.get("id", "?"))
            continue
        tenant_id   = int(export.get("tenant_id") or 0)
        tenant_name = tenant_names.get(tenant_id, str(tenant_id))
        if tenant_name not in tenant_hosts:
            log.debug("Skipping NFS export id=%s — tenant %r not in tenant_hosts", export.get("id", "?"), tenant_name)
            continue
        data_host = _tenant_fqdn(tenant_id, tenant_names, tenant_hosts)
        export_path = export.get("export_path", "")
        if not export_path:
            log.warning("NFS export id=%s has no export_path — path will be incomplete", export.get("id", "?"))

        quota = quota_map.get(fs_path)
        limit = int(quota.get("limit") or 0)          if quota is not None else None
        used  = int(quota.get("capacity_usage") or 0) if quota is not None else None

        metadata = _metadata(fs_path)

        rows.append({
            "cluster_name": qumulo.cluster_name,
            "share_path":   f"{data_host}:{export_path}",
            "type":         "NFS",
            "fs_path":      fs_path,
            "quota_limit":  _bytes_to_human(limit) if limit is not None else "",
            "quota_used":   _bytes_to_human(used)  if used  is not None else "",
            "owner":        metadata.get("Owner", ""),
            "team":         metadata.get("Team", ""),
            "snow":         metadata.get("SNOW Request", ""),
            "speedtype":    metadata.get("Speedtype", ""),
            "tenant_id":    tenant_id,
            "tenant_name":  tenant_names.get(tenant_id, str(tenant_id)),
        })

    return rows

# ---------------------------------------------------------------------------
# Excel report builder
# ---------------------------------------------------------------------------

# Colours
_BLUE_FILL   = PatternFill("solid", fgColor="4472C4")   # header row
_THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"),  bottom=Side(style="thin"),
)

# Column layout: (header label, attribute key, width)
_COLUMNS = [
    ("Share Path/NFS Export", "share_path",  60),
    ("SMB/NFS",               "type",        12),
    ("Filesystem Path",       "fs_path",     32),
    ("Quota Allocated",       "quota_limit", 18),
    ("Quota Used",            "quota_used",  18),
    ("Owner",                 "owner",       22),
    ("Team",                  "team",        22),
    ("SNOW Request",          "snow",        20),
    ("Speedtype",             "speedtype",   16),
]


def _excel_sheet_name(name: str) -> str:
    """Return *name* sanitised for use as an Excel worksheet name.

    Excel sheet names must be <= 31 characters, must not contain the
    characters \\ / ? * [ ] :, and must not start or end with a single quote.
    """
    for ch in r"\/?*[]":
        name = name.replace(ch, "_")
    name = name.replace(":", "_")
    name = name.strip("'")        # Excel rejects names that start or end with '
    return name[:31] or "unnamed"  # guard against a name that was all apostrophes


def _build_sheet(ws, rows: list[dict]) -> None:
    """Populate *ws* with a header row and one data row per entry in *rows*."""
    for col_idx, (header, _, width) in enumerate(_COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill      = _BLUE_FILL
        cell.font      = Font(bold=True, color="FFFFFF")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border    = _THIN_BORDER
        ws.column_dimensions[cell.column_letter].width = width

    for row_idx, row in enumerate(rows, start=2):
        for col_idx, (_, key, _) in enumerate(_COLUMNS, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=row.get(key, ""))
            cell.alignment = Alignment(vertical="center")
            cell.border    = _THIN_BORDER


def _unique_sheet_name(base: str, used: set[str]) -> str:
    """Return *base* (already sanitised) with a numeric suffix if it collides."""
    if base not in used:
        return base
    for n in range(2, 10_000):
        suffix = f"_{n}"
        candidate = base[: 31 - len(suffix)] + suffix
        if candidate not in used:
            return candidate
    raise ValueError(f"Cannot generate a unique sheet name from '{base}'")


def build_excel_report(cluster_rows: dict[str, list[dict]]) -> bytes:
    """Build a workbook with one sheet per cluster, each named after the cluster.

    *cluster_rows* maps cluster_name → list of row dicts.  If the dict is
    empty (all clusters failed) the workbook contains a single blank sheet
    so that the file remains valid.
    """
    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # remove the default empty sheet

    if not cluster_rows:
        ws = wb.create_sheet(title="No Data")
        _build_sheet(ws, [])
    else:
        used_names: set[str] = set()
        for cluster_name, rows in cluster_rows.items():
            title = _unique_sheet_name(_excel_sheet_name(cluster_name), used_names)
            used_names.add(title)
            ws = wb.create_sheet(title=title)
            _build_sheet(ws, rows)

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()

# ---------------------------------------------------------------------------
# Mailer
# ---------------------------------------------------------------------------

class ReportMailer:
    def __init__(self, cfg: SmtpConfig) -> None:
        self.cfg = cfg

    def send(self, subject: str, body: str, attachment: bytes, filename: str) -> None:
        msg = EmailMessage()
        msg["Subject"] = subject
        msg["From"]    = self.cfg.from_addr
        msg["To"]      = ", ".join(self.cfg.to_addrs)
        msg.set_content(body)
        msg.add_attachment(
            attachment,
            maintype="application",
            subtype="vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            filename=filename,
        )

        server = smtplib.SMTP(self.cfg.host, self.cfg.port, timeout=30)
        try:
            if self.cfg.use_tls:
                server.starttls()
            if self.cfg.username and self.cfg.password:
                server.login(self.cfg.username, self.cfg.password)
            server.send_message(msg)
            log.info("Report emailed to %s", self.cfg.to_addrs)
        finally:
            try:
                server.quit()
            except Exception:
                pass  # ignore quit errors — the send result (success or exception) is what matters

# ---------------------------------------------------------------------------
# Core collection + reporting logic
# ---------------------------------------------------------------------------

def _fetch_cluster_data(cluster: QumuloConfig) -> tuple:
    """
    Connect to a single Qumulo cluster and fetch all data needed for metrics
    and the report.

    Returns (qumulo, tenant_names, quotas, shares, exports, rows).
    """
    qumulo       = QumuloClient(cluster)
    tenant_names = qumulo.get_tenants()
    quotas       = qumulo.get_quotas()
    shares       = qumulo.get_smb_shares()
    exports      = qumulo.get_nfs_exports()
    rows         = assemble_report_rows(
        qumulo=qumulo,
        tenant_names=tenant_names,
        tenant_hosts=cluster.tenant_hosts,
        smb_shares=shares,
        nfs_exports=exports,
        quotas=quotas,
    )
    return qumulo, tenant_names, quotas, shares, exports, rows


def collect_and_store(cfg: AppConfig, dry_run: bool = False,
                      _data: list[tuple | None] | None = None) -> list[tuple | None]:
    """
    Pull data from every configured cluster and write metrics to InfluxDB.

    *_data* is a list with one entry per cluster (aligned with cfg.clusters).
    If supplied, the API fetch is skipped and the pre-fetched data is used —
    avoiding a redundant round-trip when the monthly report is sent on the
    same cycle.  A None entry means that cluster's fetch already failed.

    Returns the per-cluster data list so the caller can reuse it.
    """
    log.info("Starting data collection across %d cluster(s)%s",
             len(cfg.clusters), "  [DRY RUN]" if dry_run else "")

    if _data is None:
        _data = []
        for cluster in cfg.clusters:
            try:
                _data.append(_fetch_cluster_data(cluster))
            except Exception:
                log.exception("Failed to fetch data from cluster '%s'", cluster.cluster_name)
                _data.append(None)

    for cluster, cluster_data in zip(cfg.clusters, _data):
        if cluster_data is None:
            log.warning("Skipping InfluxDB write for cluster '%s' — data unavailable",
                        cluster.cluster_name)
            continue
        _qumulo, tenant_names, quotas, shares, exports, rows = cluster_data
        log.info("  Cluster '%s': %d quotas, %d SMB shares, %d NFS exports",
                 cluster.cluster_name, len(quotas), len(shares), len(exports))
        if dry_run:
            log.info("  DRY RUN — InfluxDB write skipped for '%s'", cluster.cluster_name)
        else:
            try:
                influx = InfluxWriter(cfg.influxdb, cluster.cluster_name)
                try:
                    influx.write_quota_metrics(quotas)
                    influx.write_smb_share_count(shares, tenant_names)
                    influx.write_nfs_export_count(exports, tenant_names)
                    influx.write_share_details(rows)
                finally:
                    influx.close()
            except Exception:
                log.exception("Failed to write InfluxDB data for cluster '%s'", cluster.cluster_name)

    log.info("Data collection complete")
    return _data


def send_monthly_report(cfg: AppConfig, dry_run: bool = False,
                        _data: list[tuple | None] | None = None) -> None:
    """
    Build an Excel report from all configured clusters and email it.

    *_data* is a per-cluster list (aligned with cfg.clusters).  If not
    supplied, data is fetched fresh.  None entries (failed clusters) are
    skipped; each remaining cluster gets its own sheet named after the cluster.
    """
    log.info("Building monthly report across %d cluster(s)%s",
             len(cfg.clusters), "  [DRY RUN]" if dry_run else "")

    if _data is None:
        _data = []
        for cluster in cfg.clusters:
            try:
                _data.append(_fetch_cluster_data(cluster))
            except Exception:
                log.exception("Failed to fetch data from cluster '%s' for report",
                              cluster.cluster_name)
                _data.append(None)

    cluster_rows: dict[str, list[dict]] = {}
    for cluster, cluster_data in zip(cfg.clusters, _data):
        if cluster_data is None:
            log.warning("Omitting cluster '%s' from report — data unavailable",
                        cluster.cluster_name)
            continue
        _qumulo, _tenant_names, _quotas, _shares, _exports, rows = cluster_data
        cluster_rows[cluster.cluster_name] = rows

    if not cluster_rows:
        log.error("No cluster data available — report not sent")
        return

    xlsx = build_excel_report(cluster_rows)

    # Use local time to match the scheduler in run_service, which also uses
    # local time (datetime.now()).  Using UTC here could produce a filename for
    # the wrong month when the service is in a timezone offset from UTC and
    # the report fires near a month boundary.
    now      = datetime.now()
    filename = f"storage_report_{now.strftime('%Y-%m')}.xlsx"

    if dry_run:
        out_path = Path(__file__).parent / filename
        out_path.write_bytes(xlsx)
        log.info("DRY RUN — email skipped.  Excel saved to: %s", out_path)
    else:
        mailer = ReportMailer(cfg.smtp)
        mailer.send(
            subject=f"{cfg.report.title} – {now.strftime('%B %Y')}",
            body=f"Please find attached the {cfg.report.title} for {now.strftime('%B %Y')}.",
            attachment=xlsx,
            filename=filename,
        )

# ---------------------------------------------------------------------------
# Scheduler / service loop
# ---------------------------------------------------------------------------

_running = True


def _load_last_report_ym() -> int | None:
    """Return the YYYYMM integer of the last successfully sent report.

    Reads from *_STATE_PATH*.  Returns None if the file does not exist or
    cannot be parsed — which causes the service to treat no report as having
    been sent yet (safe for a fresh install, re-sends on same-day restart if
    the file was lost, but that is the correct fallback).
    """
    try:
        return int(_STATE_PATH.read_text(encoding="utf-8").strip())
    except (OSError, ValueError):
        return None


def _save_last_report_ym(ym: int) -> None:
    """Persist *ym* (YYYYMM) so a restarted service does not re-send the report."""
    try:
        _STATE_PATH.write_text(str(ym), encoding="utf-8")
    except OSError:
        log.warning(
            "Could not write report state to %s — a restart on the same day "
            "may send the report again", _STATE_PATH,
        )


def _handle_signal(signum, frame):  # noqa: ANN001
    global _running
    log.info("Received signal %s – shutting down", signum)
    _running = False


def run_service(cfg: AppConfig, collection_interval_seconds: int = 300,
                dry_run: bool = False) -> None:
    """
    Main service loop.
    - Collects metrics from Qumulo every `collection_interval_seconds`.
    - Sends the monthly Excel report on the configured day/hour.
    """
    global _running
    _running = True

    signal.signal(signal.SIGINT, _handle_signal)
    if sys.platform != "win32":
        signal.signal(signal.SIGTERM, _handle_signal)

    # YYYYMM of the last successfully sent report.  Loaded from disk so that
    # a restart on the same scheduled day does not re-send the report.
    last_report_ym = _load_last_report_ym()

    log.info(
        "Service started%s. Clusters: %s. Collection every %ds. "
        "Monthly report on day %d at %02d:00.",
        " [DRY RUN]" if dry_run else "",
        [c.cluster_name for c in cfg.clusters],
        collection_interval_seconds,
        cfg.report.schedule_day,
        cfg.report.schedule_hour,
    )

    while _running:
        fetched = None
        try:
            fetched = collect_and_store(cfg, dry_run=dry_run)
        except Exception:
            log.exception("Collection failed")

        # Capture time after collection so the hour check reflects when
        # collection actually finished, not when it started.
        now    = datetime.now()
        now_ym = now.year * 100 + now.month
        if (
            now.day  == cfg.report.schedule_day
            and now.hour >= cfg.report.schedule_hour
            and last_report_ym != now_ym
        ):
            try:
                send_monthly_report(cfg, dry_run=dry_run, _data=fetched)
                last_report_ym = now_ym
                _save_last_report_ym(now_ym)
            except Exception:
                log.exception("Failed to send monthly report")

        for _ in range(collection_interval_seconds):
            if not _running:
                break
            time.sleep(1)

    log.info("Service stopped.")

# ---------------------------------------------------------------------------
# Unit tests  (run with: python storage_report.py --test)
# ---------------------------------------------------------------------------

def _b64(s: str) -> str:
    return base64.b64encode(s.encode()).decode()

_MOCK_TENANTS = {
    "entries": [
        {"id": 1, "name": "Default",        "smb_enabled": True, "nfs_enabled": True},
        {"id": 2, "name": "Finance_Tenant", "smb_enabled": True, "nfs_enabled": False},
    ]
}
_MOCK_NFS_EXPORTS = {
    "entries": [
        {"id": "1",  "export_path": "/nfs_data",     "tenant_id": 1, "fs_path": "/nfs_server_1/"},
        {"id": "2",  "export_path": "/data_export",  "tenant_id": 1, "fs_path": "/data/"},
        {"id": "3",  "export_path": "/finance_export","tenant_id": 2, "fs_path": "/Finance/"},
        {"id": "10", "export_path": "/zzz-export",   "tenant_id": 1, "fs_path": "/zzz/"},
    ]
}
_MOCK_SMB_SHARES = {
    "entries": [
        {"id": "3",  "share_name": "data",         "tenant_id": 1, "fs_path": "/data"},
        {"id": "11", "share_name": "zzz",          "tenant_id": 1, "fs_path": "/zzz/"},
        {"id": "22", "share_name": "quota_test",   "tenant_id": 1, "fs_path": "/quota_test"},
        {"id": "7",  "share_name": "finance_share","tenant_id": 2, "fs_path": "/share_server2"},
    ]
}
_MOCK_QUOTAS = {
    "quotas": [
        {"id": "1", "path": "/zzz/",         "limit": "31000000000",  "capacity_usage": "27293859840"},
        {"id": "2", "path": "/data/",         "limit": "100000000000", "capacity_usage": "32212815872"},
        {"id": "3", "path": "/quota_test/",   "limit": "10000000000",  "capacity_usage": "7993434112"},
        {"id": "4", "path": "/nfs_server_1/", "limit": "0",            "capacity_usage": "0"},
    ],
    "paging": {"next": ""},
}
_MOCK_FILE_METADATA = {
    "entries": [
        {"type": "GENERIC", "key": "Owner",        "value": _b64("Andy Kress")},
        {"type": "GENERIC", "key": "Team",         "value": _b64("Platform")},
        {"type": "GENERIC", "key": "SNOW Request", "value": _b64("RITM00000000")},
        {"type": "GENERIC", "key": "Speedtype",    "value": _b64("123456")},
    ],
    "paging": {"next": None},
}
_MOCK_TENANT_HOSTS = {
    "Default":        "nas01.example.com",
    "Finance_Tenant": "nas02.example.com",
}


def _mock_qumulo_client() -> "QumuloClient":
    cfg    = QumuloConfig(host="testhost", token="testtoken", cluster_name="TestCluster")
    client = QumuloClient(cfg)

    def fake_get(path: str, params=None):
        if "/multitenancy/tenants" in path: return _MOCK_TENANTS
        if "/nfs/exports"         in path: return _MOCK_NFS_EXPORTS
        if "/smb/shares"          in path: return _MOCK_SMB_SHARES
        if "/quotas/status"       in path: return _MOCK_QUOTAS
        if "/user-metadata/generic" in path: return _MOCK_FILE_METADATA
        raise ValueError(f"Unexpected API path in test: {path}")

    client._get = fake_get
    return client


class _TestQumuloClient(unittest.TestCase):

    def setUp(self):
        self.client = _mock_qumulo_client()

    def test_get_tenants(self):
        self.assertEqual(self.client.get_tenants(), {1: "Default", 2: "Finance_Tenant"})

    def test_get_nfs_exports(self):
        result = self.client.get_nfs_exports()
        self.assertEqual(len(result), 4)
        self.assertEqual(result[0]["export_path"], "/nfs_data")

    def test_get_smb_shares(self):
        result = self.client.get_smb_shares()
        self.assertEqual(len(result), 4)
        self.assertEqual(result[0]["share_name"], "data")

    def test_get_quotas(self):
        result = self.client.get_quotas()
        self.assertEqual(len(result), 4)
        self.assertEqual(result[0]["path"], "/zzz/")
        self.assertEqual(int(result[0]["limit"]), 31_000_000_000)

    def test_get_file_metadata_decodes_base64(self):
        result = self.client.get_file_metadata("/data")
        self.assertEqual(result["Owner"],        "Andy Kress")
        self.assertEqual(result["Team"],         "Platform")
        self.assertEqual(result["SNOW Request"], "RITM00000000")
        self.assertEqual(result["Speedtype"],    "123456")

    def test_get_file_metadata_returns_empty_on_http_error(self):
        cfg    = QumuloConfig(host="h", token="t", cluster_name="c")
        client = QumuloClient(cfg)
        fake_resp = MagicMock()
        fake_resp.status_code = 404
        client._get = MagicMock(side_effect=requests.HTTPError(response=fake_resp))
        self.assertEqual(client.get_file_metadata("/no/such/path"), {})

    def test_get_file_metadata_returns_empty_on_connection_error(self):
        cfg    = QumuloConfig(host="h", token="t", cluster_name="c")
        client = QumuloClient(cfg)
        client._get = MagicMock(side_effect=requests.ConnectionError("timed out"))
        self.assertEqual(client.get_file_metadata("/some/path"), {})

    def test_get_file_metadata_falls_back_to_raw_on_bad_base64(self):
        cfg    = QumuloConfig(host="h", token="t", cluster_name="c")
        client = QumuloClient(cfg)
        client._get = MagicMock(return_value={
            "entries": [{"key": "Owner", "value": "!!!not-valid-base64!!!"}],
            "paging":  {"next": None},
        })
        result = client.get_file_metadata("/some/path")
        # Should fall back to raw value rather than raising
        self.assertEqual(result["Owner"], "!!!not-valid-base64!!!")


class _TestBytesToHuman(unittest.TestCase):

    def test_petabytes(self):  self.assertEqual(_bytes_to_human(2_000_000_000_000_000), "2.00PB")
    def test_terabytes(self): self.assertEqual(_bytes_to_human(1_600_000_000_000),     "1.60TB")
    def test_gigabytes(self):  self.assertEqual(_bytes_to_human(32_212_815_872),       "32.21GB")
    def test_megabytes(self):  self.assertEqual(_bytes_to_human(500_000_000),           "500.00MB")
    def test_bytes(self):      self.assertEqual(_bytes_to_human(999),                   "999B")
    def test_zero(self):       self.assertEqual(_bytes_to_human(0),                     "0B")
    def test_negative(self):   self.assertEqual(_bytes_to_human(-100),                  "0B")


class _TestTenantFqdn(unittest.TestCase):

    def test_known_tenant(self):
        self.assertEqual(_tenant_fqdn(2, {1: "Default", 2: "Finance_Tenant"}, _MOCK_TENANT_HOSTS),
                         "nas02.example.com")

    def test_unknown_tenant_fallback(self):
        self.assertEqual(_tenant_fqdn(99, {1: "Default"}, _MOCK_TENANT_HOSTS), "unknown-tenant")


class _TestAssembleReportRows(unittest.TestCase):

    def setUp(self):
        client = _mock_qumulo_client()
        self.tenant_names = {1: "Default", 2: "Finance_Tenant"}
        self.rows = assemble_report_rows(
            qumulo=client,
            tenant_names=self.tenant_names,
            tenant_hosts=_MOCK_TENANT_HOSTS,
            smb_shares=_MOCK_SMB_SHARES["entries"],
            nfs_exports=_MOCK_NFS_EXPORTS["entries"],
            quotas=_MOCK_QUOTAS["quotas"],
        )

    def test_row_count(self):
        self.assertEqual(len(self.rows), 8)

    def test_smb_share_path_format(self):
        row = next(r for r in self.rows if r["type"] == "SMB" and r["fs_path"] == "/data")
        self.assertEqual(row["share_path"], r"\\nas01.example.com\data")

    def test_nfs_export_path_format(self):
        row = next(r for r in self.rows if r["type"] == "NFS" and r["fs_path"] == "/zzz")
        self.assertEqual(row["share_path"], "nas01.example.com:/zzz-export")

    def test_quota_matched_by_fs_path(self):
        row = next(r for r in self.rows if r["type"] == "SMB" and r["fs_path"] == "/zzz")
        self.assertEqual(row["quota_limit"], "31.00GB")
        self.assertEqual(row["quota_used"],  "27.29GB")

    def test_no_quota_leaves_empty_string(self):
        row = next(r for r in self.rows if r["fs_path"] == "/Finance")
        self.assertEqual(row["quota_limit"], "")
        self.assertEqual(row["quota_used"],  "")

    def test_quota_zero_limit_displays(self):
        row = next(r for r in self.rows if r["fs_path"] == "/nfs_server_1")
        self.assertEqual(row["quota_limit"], "0B")
        self.assertEqual(row["quota_used"],  "0B")

    def test_nfs_fs_path_correct(self):
        for row in (r for r in self.rows if r["type"] == "NFS"):
            self.assertNotEqual(row["fs_path"], "/", f"NFS row has fs_path '/' — wrong field used")

    def test_file_metadata_populated(self):
        row = next(r for r in self.rows if r["type"] == "SMB" and r["fs_path"] == "/data")
        self.assertEqual(row["owner"], "Andy Kress")
        self.assertEqual(row["team"],  "Platform")
        self.assertEqual(row["snow"],  "RITM00000000")

    def test_tenant_name_resolved(self):
        row = next(r for r in self.rows if r["tenant_id"] == 2)
        self.assertEqual(row["tenant_name"], "Finance_Tenant")

    def test_trailing_slash_normalisation(self):
        row = next(r for r in self.rows if r["type"] == "SMB" and r["fs_path"] == "/zzz")
        self.assertFalse(row["fs_path"].endswith("/"))

    def test_empty_share_name_row_still_emitted(self):
        client = _mock_qumulo_client()
        rows = assemble_report_rows(
            qumulo=client,
            tenant_names={1: "Default"},
            tenant_hosts=_MOCK_TENANT_HOSTS,
            smb_shares=[{"id": "99", "share_name": "", "tenant_id": 1, "fs_path": "/data"}],
            nfs_exports=[],
            quotas=[],
        )
        self.assertEqual(len(rows), 1)
        self.assertEqual(rows[0]["share_path"], "\\\\nas01.example.com\\")

    def test_empty_export_path_row_still_emitted(self):
        client = _mock_qumulo_client()
        rows = assemble_report_rows(
            qumulo=client,
            tenant_names={1: "Default"},
            tenant_hosts=_MOCK_TENANT_HOSTS,
            smb_shares=[],
            nfs_exports=[{"id": "99", "export_path": "", "tenant_id": 1, "fs_path": "/data"}],
            quotas=[],
        )
        self.assertEqual(len(rows), 1)
        self.assertEqual(rows[0]["share_path"], "nas01.example.com:")

    def test_smb_root_fs_path_is_skipped(self):
        client = _mock_qumulo_client()
        rows = assemble_report_rows(
            qumulo=client,
            tenant_names={1: "Default"},
            tenant_hosts=_MOCK_TENANT_HOSTS,
            smb_shares=[
                {"id": "1", "share_name": "root_share", "tenant_id": 1, "fs_path": "/"},
                {"id": "2", "share_name": "data",       "tenant_id": 1, "fs_path": "/data"},
            ],
            nfs_exports=[],
            quotas=[],
        )
        self.assertEqual(len(rows), 1)
        self.assertEqual(rows[0]["fs_path"], "/data")

    def test_nfs_root_fs_path_is_skipped(self):
        client = _mock_qumulo_client()
        rows = assemble_report_rows(
            qumulo=client,
            tenant_names={1: "Default"},
            tenant_hosts=_MOCK_TENANT_HOSTS,
            smb_shares=[],
            nfs_exports=[
                {"id": "1", "export_path": "/",     "tenant_id": 1, "fs_path": "/"},
                {"id": "2", "export_path": "/data", "tenant_id": 1, "fs_path": "/data"},
            ],
            quotas=[],
        )
        self.assertEqual(len(rows), 1)
        self.assertEqual(rows[0]["fs_path"], "/data")

    def test_smb_unknown_tenant_is_skipped(self):
        client = _mock_qumulo_client()
        rows = assemble_report_rows(
            qumulo=client,
            tenant_names={1: "Default", 99: "MGMT"},
            tenant_hosts=_MOCK_TENANT_HOSTS,  # only has "Default" and "Finance_Tenant"
            smb_shares=[
                {"id": "1", "share_name": "mgmt_share", "tenant_id": 99, "fs_path": "/mgmt"},
                {"id": "2", "share_name": "data",       "tenant_id": 1,  "fs_path": "/data"},
            ],
            nfs_exports=[],
            quotas=[],
        )
        self.assertEqual(len(rows), 1)
        self.assertEqual(rows[0]["fs_path"], "/data")

    def test_nfs_unknown_tenant_is_skipped(self):
        client = _mock_qumulo_client()
        rows = assemble_report_rows(
            qumulo=client,
            tenant_names={1: "Default", 99: "MGMT"},
            tenant_hosts=_MOCK_TENANT_HOSTS,  # only has "Default" and "Finance_Tenant"
            smb_shares=[],
            nfs_exports=[
                {"id": "1", "export_path": "/mgmt", "tenant_id": 99, "fs_path": "/mgmt"},
                {"id": "2", "export_path": "/data", "tenant_id": 1,  "fs_path": "/data"},
            ],
            quotas=[],
        )
        self.assertEqual(len(rows), 1)
        self.assertEqual(rows[0]["fs_path"], "/data")


class _TestBuildExcelReport(unittest.TestCase):

    def setUp(self):
        client = _mock_qumulo_client()
        rows   = assemble_report_rows(
            qumulo=client,
            tenant_names={1: "Default", 2: "Finance_Tenant"},
            tenant_hosts=_MOCK_TENANT_HOSTS,
            smb_shares=_MOCK_SMB_SHARES["entries"],
            nfs_exports=_MOCK_NFS_EXPORTS["entries"],
            quotas=_MOCK_QUOTAS["quotas"],
        )
        self.wb = openpyxl.load_workbook(BytesIO(build_excel_report({"TestCluster": rows})))
        self.ws = self.wb["TestCluster"]

    def test_single_sheet(self):
        self.assertEqual(self.wb.sheetnames, ["TestCluster"])

    def test_row1_headers(self):
        headers = [self.ws.cell(row=1, column=c).value for c in range(1, 10)]
        self.assertEqual(headers[0], "Share Path/NFS Export")
        self.assertEqual(headers[1], "SMB/NFS")
        self.assertEqual(headers[5], "Owner")

    def test_data_starts_at_row2(self):
        self.assertIsNotNone(self.ws.cell(row=2, column=1).value)

    def test_data_row_count(self):
        # 4 SMB shares + 4 NFS exports = 8 data rows; max_row includes the header row
        self.assertEqual(self.ws.max_row - 1, 8)

    def test_empty_cluster_rows_produces_no_data_sheet(self):
        wb = openpyxl.load_workbook(BytesIO(build_excel_report({})))
        self.assertEqual(wb.sheetnames, ["No Data"])

    def test_sanitized_name_collision_deduplicates(self):
        # "prod:1" and "prod_1" both sanitise to "prod_1" — the second must get a suffix
        wb = openpyxl.load_workbook(BytesIO(build_excel_report({"prod:1": [], "prod_1": []})))
        self.assertEqual(len(wb.sheetnames), 2)
        self.assertNotEqual(wb.sheetnames[0], wb.sheetnames[1])

    def test_multiple_clusters_produce_separate_sheets(self):
        client = _mock_qumulo_client()
        rows = assemble_report_rows(
            qumulo=client,
            tenant_names={1: "Default", 2: "Finance_Tenant"},
            tenant_hosts=_MOCK_TENANT_HOSTS,
            smb_shares=_MOCK_SMB_SHARES["entries"],
            nfs_exports=_MOCK_NFS_EXPORTS["entries"],
            quotas=_MOCK_QUOTAS["quotas"],
        )
        wb = openpyxl.load_workbook(BytesIO(build_excel_report({
            "ClusterA": rows[:2],
            "ClusterB": rows[2:],
        })))
        self.assertEqual(wb.sheetnames, ["ClusterA", "ClusterB"])
        # Each sheet has its own independent row count
        ws_a = wb["ClusterA"]
        ws_b = wb["ClusterB"]
        self.assertEqual(ws_a.max_row - 1, 2)   # 2 data rows
        self.assertEqual(ws_b.max_row - 1, len(rows) - 2)  # remaining rows
        # Both sheets have the same header in column 1
        self.assertEqual(ws_a.cell(row=1, column=1).value, "Share Path/NFS Export")
        self.assertEqual(ws_b.cell(row=1, column=1).value, "Share Path/NFS Export")


class _TestExcelSheetName(unittest.TestCase):

    def test_illegal_chars_replaced(self):
        self.assertEqual(_excel_sheet_name(r"a\b/c?d*e[f]g:h"), "a_b_c_d_e_f_g_h")

    def test_truncated_to_31_chars(self):
        self.assertEqual(len(_excel_sheet_name("x" * 40)), 31)

    def test_leading_trailing_apostrophe_stripped(self):
        self.assertEqual(_excel_sheet_name("'hello'"), "hello")

    def test_all_apostrophes_falls_back_to_unnamed(self):
        self.assertEqual(_excel_sheet_name("'''"), "unnamed")

    def test_normal_name_unchanged(self):
        self.assertEqual(_excel_sheet_name("prod-cluster-01"), "prod-cluster-01")


class _TestValidateConfig(unittest.TestCase):

    def _valid_cfg(self) -> AppConfig:
        return AppConfig(
            clusters=[QumuloConfig(host="h", token="t", cluster_name="c")],
            influxdb=InfluxConfig(url="u", token="t", org="o", bucket="b"),
            smtp=SmtpConfig(host="smtp", port=25, from_addr="f@x.com", to_addrs=["t@x.com"]),
            report=ReportConfig(schedule_day=1, schedule_hour=8, collection_interval_minutes=5),
        )

    def test_valid_config_passes(self):
        _validate_config(self._valid_cfg())  # must not raise

    def test_schedule_day_zero_raises(self):
        cfg = self._valid_cfg(); cfg.report.schedule_day = 0
        with self.assertRaises(ValueError): _validate_config(cfg)

    def test_schedule_day_29_raises(self):
        cfg = self._valid_cfg(); cfg.report.schedule_day = 29
        with self.assertRaises(ValueError): _validate_config(cfg)

    def test_schedule_hour_negative_raises(self):
        cfg = self._valid_cfg(); cfg.report.schedule_hour = -1
        with self.assertRaises(ValueError): _validate_config(cfg)

    def test_schedule_hour_24_raises(self):
        cfg = self._valid_cfg(); cfg.report.schedule_hour = 24
        with self.assertRaises(ValueError): _validate_config(cfg)

    def test_interval_zero_raises(self):
        cfg = self._valid_cfg(); cfg.report.collection_interval_minutes = 0
        with self.assertRaises(ValueError): _validate_config(cfg)

    def test_empty_to_addrs_raises(self):
        cfg = self._valid_cfg(); cfg.smtp.to_addrs = []
        with self.assertRaises(ValueError): _validate_config(cfg)

    def test_no_clusters_raises(self):
        cfg = self._valid_cfg(); cfg.clusters = []
        with self.assertRaises(ValueError): _validate_config(cfg)

    def test_duplicate_cluster_name_raises(self):
        cfg = self._valid_cfg()
        cfg.clusters = [
            QumuloConfig(host="h1", token="t1", cluster_name="same"),
            QumuloConfig(host="h2", token="t2", cluster_name="same"),
        ]
        with self.assertRaises(ValueError): _validate_config(cfg)

    def test_empty_cluster_name_raises(self):
        cfg = self._valid_cfg()
        cfg.clusters = [QumuloConfig(host="h", token="t", cluster_name="")]
        with self.assertRaises(ValueError): _validate_config(cfg)

    def test_partial_smtp_credentials_raises(self):
        cfg = self._valid_cfg()
        cfg.smtp.username = "user@example.com"
        cfg.smtp.password = None
        with self.assertRaises(ValueError): _validate_config(cfg)

    def test_smtp_port_zero_raises(self):
        cfg = self._valid_cfg(); cfg.smtp.port = 0
        with self.assertRaises(ValueError): _validate_config(cfg)

    def test_smtp_port_too_high_raises(self):
        cfg = self._valid_cfg(); cfg.smtp.port = 65536
        with self.assertRaises(ValueError): _validate_config(cfg)

    def test_empty_cluster_host_raises(self):
        cfg = self._valid_cfg()
        cfg.clusters = [QumuloConfig(host="", token="t", cluster_name="c")]
        with self.assertRaises(ValueError): _validate_config(cfg)

    def test_empty_cluster_token_raises(self):
        cfg = self._valid_cfg()
        cfg.clusters = [QumuloConfig(host="h", token="", cluster_name="c")]
        with self.assertRaises(ValueError): _validate_config(cfg)

    def test_empty_influxdb_url_raises(self):
        cfg = self._valid_cfg(); cfg.influxdb.url = ""
        with self.assertRaises(ValueError): _validate_config(cfg)

    def test_empty_smtp_host_raises(self):
        cfg = self._valid_cfg(); cfg.smtp.host = ""
        with self.assertRaises(ValueError): _validate_config(cfg)


class _TestLoadConfig(unittest.TestCase):

    def test_load_valid_config(self):
        toml = b"""
[[clusters]]
host         = "192.168.1.1"
port         = 8000
token        = "tok"
cluster_name = "TestCluster"
[clusters.tenant_hosts]
"Default" = "nas01.example.com"
[influxdb]
url    = "http://localhost:8086"
token  = "influx-tok"
org    = "my-org"
bucket = "Qumulo storage reporting"
[smtp]
host      = "smtp.example.com"
port      = 587
use_tls   = true
username  = "u"
password  = "p"
from_addr = "from@example.com"
to_addrs  = ["to@example.com"]
[report]
schedule_day  = 1
schedule_hour = 8
"""
        with tempfile.NamedTemporaryFile(suffix=".toml", delete=False) as f:
            f.write(toml)
            tmp = Path(f.name)
        try:
            cfg = load_config(tmp)
            self.assertEqual(len(cfg.clusters),                   1)
            self.assertEqual(cfg.clusters[0].cluster_name,        "TestCluster")
            self.assertEqual(cfg.clusters[0].tenant_hosts,        {"Default": "nas01.example.com"})
            self.assertEqual(cfg.influxdb.bucket,                 "Qumulo storage reporting")
        finally:
            tmp.unlink()

    def test_missing_required_key_raises(self):
        # [[clusters]] entry present but missing the required 'token' key.
        # All other required sections are present so the KeyError is
        # specifically for the missing cluster token, not for a missing section.
        toml = b"""
[[clusters]]
host         = "192.168.1.1"
cluster_name = "c"
[influxdb]
url    = "http://localhost:8086"
token  = "tok"
org    = "org"
bucket = "b"
[smtp]
host      = "smtp.example.com"
port      = 25
from_addr = "f@x.com"
to_addrs  = ["t@x.com"]
[report]
schedule_day  = 1
schedule_hour = 8
"""
        with tempfile.NamedTemporaryFile(suffix=".toml", delete=False) as f:
            f.write(toml)
            tmp = Path(f.name)
        try:
            with self.assertRaises(KeyError):
                load_config(tmp)
        finally:
            tmp.unlink()


class _TestReportState(unittest.TestCase):

    def _make_tmp(self, existing: bool = False) -> Path:
        """Return a path to a temp file.  If *existing* is False the file is removed."""
        fd, name = tempfile.mkstemp()
        os.close(fd)
        p = Path(name)
        if not existing:
            p.unlink()
        return p

    def _patch_state_path(self, p: Path):
        """Context manager that temporarily replaces the module-level _STATE_PATH."""
        mod = sys.modules[__name__]

        @contextlib.contextmanager
        def _ctx():
            original = mod._STATE_PATH
            mod._STATE_PATH = p
            try:
                yield
            finally:
                mod._STATE_PATH = original

        return _ctx()

    def test_load_returns_none_when_file_missing(self):
        p = self._make_tmp(existing=False)
        with self._patch_state_path(p):
            self.assertIsNone(_load_last_report_ym())

    def test_save_then_load_roundtrip(self):
        p = self._make_tmp(existing=False)
        try:
            with self._patch_state_path(p):
                _save_last_report_ym(202601)
                self.assertEqual(_load_last_report_ym(), 202601)
        finally:
            p.unlink(missing_ok=True)

    def test_load_returns_none_on_corrupt_file(self):
        p = self._make_tmp(existing=True)
        p.write_text("not-a-number", encoding="utf-8")
        try:
            with self._patch_state_path(p):
                self.assertIsNone(_load_last_report_ym())
        finally:
            p.unlink(missing_ok=True)


def _run_unit_tests() -> None:
    suite = unittest.TestLoader().loadTestsFromTestCase(_TestQumuloClient)
    for cls in (_TestBytesToHuman, _TestTenantFqdn, _TestAssembleReportRows,
                _TestBuildExcelReport, _TestExcelSheetName,
                _TestValidateConfig, _TestLoadConfig, _TestReportState):
        suite.addTests(unittest.TestLoader().loadTestsFromTestCase(cls))
    result = unittest.TextTestRunner(verbosity=2).run(suite)
    sys.exit(0 if result.wasSuccessful() else 1)


# ---------------------------------------------------------------------------
# InfluxDB integration test  (run with: python storage_report.py --test-influx)
# ---------------------------------------------------------------------------

def _create_bucket_if_missing(client: InfluxDBClient, org_id: str, bucket_name: str) -> None:
    api      = client.buckets_api()
    existing = api.find_bucket_by_name(bucket_name)
    if existing:
        log.info("Test bucket '%s' already exists — reusing it", bucket_name)
    else:
        api.create_bucket(bucket_name=bucket_name, org_id=org_id)
        log.info("Created test bucket '%s'", bucket_name)


def _delete_test_bucket(cfg: AppConfig, bucket_name: str) -> None:
    client = InfluxDBClient(url=cfg.influxdb.url, token=cfg.influxdb.token, org=cfg.influxdb.org)
    try:
        api    = client.buckets_api()
        bucket = api.find_bucket_by_name(bucket_name)
        if bucket:
            api.delete_bucket(bucket)
            log.info("Deleted test bucket '%s'", bucket_name)
        else:
            log.warning("Bucket '%s' not found — nothing to delete", bucket_name)
    finally:
        client.close()


def _run_influx_integration_test(cfg: AppConfig, test_bucket: str) -> None:
    log.info("=== Qumulo → InfluxDB Integration Test ===")
    log.info("Clusters   : %d", len(cfg.clusters))
    log.info("InfluxDB   : %s", cfg.influxdb.url)
    log.info("Test bucket: %s", test_bucket)

    influx_client = InfluxDBClient(
        url=cfg.influxdb.url, token=cfg.influxdb.token, org=cfg.influxdb.org,
    )
    try:
        _create_bucket_if_missing(influx_client, cfg.influxdb.org, test_bucket)
    finally:
        influx_client.close()

    test_cfg = InfluxConfig(
        url=cfg.influxdb.url, token=cfg.influxdb.token,
        org=cfg.influxdb.org, bucket=test_bucket,
    )

    for cluster in cfg.clusters:
        log.info("--- Cluster: %s (%s) ---", cluster.cluster_name, cluster.host)
        try:
            qumulo       = QumuloClient(cluster)
            tenant_names = qumulo.get_tenants()
            log.info("Tenants    : %s", tenant_names)
            quotas  = qumulo.get_quotas()
            log.info("Quotas     : %d entries", len(quotas))
            shares  = qumulo.get_smb_shares()
            log.info("SMB shares : %d entries", len(shares))
            exports = qumulo.get_nfs_exports()
            log.info("NFS exports: %d entries", len(exports))

            rows = assemble_report_rows(
                qumulo=qumulo,
                tenant_names=tenant_names,
                tenant_hosts=cluster.tenant_hosts,
                smb_shares=shares,
                nfs_exports=exports,
                quotas=quotas,
            )
            log.info("Assembled  : %d share/export rows", len(rows))

            writer = InfluxWriter(test_cfg, cluster.cluster_name)
            try:
                writer.write_quota_metrics(quotas)
                writer.write_smb_share_count(shares, tenant_names)
                writer.write_nfs_export_count(exports, tenant_names)
                writer.write_share_details(rows)
            finally:
                writer.close()
        except Exception:
            log.exception("Integration test failed for cluster '%s' — continuing with remaining clusters",
                          cluster.cluster_name)

    log.info("=== Done. Inspect bucket '%s' in InfluxDB UI ===", test_bucket)
    log.info("When finished: python storage_report.py --test-influx-cleanup")


# ---------------------------------------------------------------------------
# Entry point
# ---------------------------------------------------------------------------

if __name__ == "__main__":
    logging.basicConfig(
        level=logging.INFO,
        format="%(asctime)s  %(levelname)-8s  %(message)s",
        datefmt="%Y-%m-%d %H:%M:%S",
    )

    parser = argparse.ArgumentParser(description="Qumulo storage reporting service")
    parser.add_argument("--dry-run",      action="store_true",
        help="Skip InfluxDB writes; save Excel to disk instead of emailing.")
    parser.add_argument("--send-report",  action="store_true",
        help="Build and send the report immediately then exit.")
    parser.add_argument("--test",         action="store_true",
        help="Run unit tests (no cluster or InfluxDB required).")
    parser.add_argument("--test-influx",  action="store_true",
        help="Run InfluxDB integration test against the real cluster and InfluxDB.")
    parser.add_argument("--test-influx-cleanup", action="store_true",
        help="Delete the InfluxDB test bucket.")
    parser.add_argument("--test-bucket",  default="test-qumulo-reporting",
        help="Test bucket name used by --test-influx / --test-influx-cleanup (default: test-qumulo-reporting).")
    parser.add_argument("--config", type=Path, default=CONFIG_PATH,
        help=f"Path to config file (default: {CONFIG_PATH})")
    args = parser.parse_args()

    # Unit tests don't need a config file
    if args.test:
        _run_unit_tests()

    # Everything else requires config
    try:
        config = load_config(args.config)
    except Exception as exc:
        log.error("Failed to load config: %s", exc)
        sys.exit(1)

    if args.test_influx:
        _run_influx_integration_test(config, args.test_bucket)
    elif args.test_influx_cleanup:
        _delete_test_bucket(config, args.test_bucket)
    elif args.send_report:
        send_monthly_report(config, dry_run=args.dry_run)
    else:
        run_service(
            config,
            collection_interval_seconds=config.report.collection_interval_minutes * 60,
            dry_run=args.dry_run,
        )
